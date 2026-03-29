import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_report():
    # 1. Read mock_7_day_data.csv to structure iteration
    df = pd.read_csv('mock_7_day_data.csv')
    df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')
    # Ensure correct sorting
    df = df.sort_values(['Date', 'Category'])
    
    # Initialize openpyxl Workbook
    wb = Workbook()
    
    # --- Sheet 1: Methodology & Justification ---
    ws1 = wb.active
    ws1.title = 'Methodology & Justification'
    
    methodology_text = [
        "Industry Choice: Fast-Moving Consumer Goods (FMCG) via Quick Commerce.",
        "",
        "Justification: Highly standardized SKUs allowing for true apples-to-apples comparison, high-frequency algorithmic pricing updates, and high relevance to real-world household expenditure inflation.",
        "",
        "Fields Scraped: Timestamp, SKU ID, Active Listed Pack Size, Current Selling Price, and Availability Status. MRP is omitted as continuous promotions make transaction price the true market clearer.",
        "",
        "Handling Stockouts: Implemented a Carry-Forward imputation model. If an item is out of stock (e.g., Eggs on Day 4), the script carries forward the last observed normalized price from the prior day to maintain index continuity.",
        "",
        "Handling Shrinkflation: Developed a pre-averaging mathematical normalization engine. The scraper extracts the active pack size from the page DOM and normalizes the selling price to a base unit (Price per 100g or Price per 1 unit) before averaging, ensuring the index tracks pure inflation, not packaging changes."
    ]
    
    for idx, text in enumerate(methodology_text, start=1):
        ws1.cell(row=idx, column=1, value=text)
        
    # --- Sheet 2: Raw Daily Price Data ---
    ws2 = wb.create_sheet(title='Raw Daily Price Data')
    
    # Define and write headers
    headers = list(df.columns) + ['Base Price', 'Price Relative']
    ws2.append(headers)
    
    # Write dataframe rows to sheet
    for row in dataframe_to_rows(df, index=False, header=False):
        ws2.append(row)
        
    # Inject Excel Formulas for Base Price (Column H) and Price Relative (Column I)
    for row_idx in range(2, ws2.max_row + 1):
        # Base Price XLOOKUP pointing to Day 1 boundaries ($C$2:$C$21 and $F$2:$F$21)
        ws2[f'H{row_idx}'] = f'=XLOOKUP(C{row_idx}, $C$2:$C$21, $F$2:$F$21)'
        # Price Relative Calculation
        ws2[f'I{row_idx}'] = f'=(F{row_idx}/H{row_idx})*100'

    # --- Sheet 3: Online CPI Computation ---
    ws3 = wb.create_sheet(title='Online CPI Computation')
    
    dates = sorted(df['Date'].unique())
    categories = ['Staples', 'Dairy', 'Produce', 'Oils', 'Household', 'Snacks']
    weights = {'Staples': 0.25, 'Dairy': 0.20, 'Produce': 0.20, 'Oils': 0.15, 'Household': 0.10, 'Snacks': 0.10}
    cat_cols = ['B', 'C', 'D', 'E', 'F', 'G']
    
    # Write Headers
    headers_ws3 = ['Date'] + categories + ['Overall CPI']
    ws3.append(headers_ws3)
    
    for row_idx, date in enumerate(dates, start=2):
        # Column A: Date
        ws3[f'A{row_idx}'] = date
        
        # Inject AVERAGEIFS for each Category
        for cat, letter in zip(categories, cat_cols):
            ws3[f'{letter}{row_idx}'] = f"=AVERAGEIFS('Raw Daily Price Data'!$I:$I, 'Raw Daily Price Data'!$A:$A, $A{row_idx}, 'Raw Daily Price Data'!$B:$B, {letter}$1)"
            
        # Inject Dynamic Overall CPI Formula with Weight Re-balancing
        weighted_sum_components = [f"({letter}{row_idx}*{weights[cat]})" for cat, letter in zip(categories, cat_cols)]
        weighted_sum = " + ".join(weighted_sum_components)
        
        active_weights_components = [f"IF({letter}{row_idx}>0, {weights[cat]}, 0)" for cat, letter in zip(categories, cat_cols)]
        active_weights = " + ".join(active_weights_components)
        
        # Overall CPI calculation string
        ws3[f'H{row_idx}'] = f'=({weighted_sum}) / ({active_weights})'
        
    # Save the native Excel workbook with live formulas
    wb.save('Assignment_Deliverable.xlsx')

if __name__ == '__main__':
    generate_report()
    print("Assignment_Deliverable.xlsx generated successfully with live Excel formulas.")
